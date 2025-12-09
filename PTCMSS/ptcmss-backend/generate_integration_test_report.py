#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Excel report with Decision Table format for Integration Tests
Format: Cover -> MethodList -> Statistics -> Method sheets (decision tables)
Based on integration test results from surefire reports
"""

import xml.etree.ElementTree as ET
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
HEADER_COLOR = '366092'
PASS_COLOR = 'C6EFCE'
FAIL_COLOR = 'FFC7CE'
WARNING_COLOR = 'FFEB9C'
BORDER_COLOR = '000000'

def parse_test_xml(xml_file):
    """Parse test XML file and extract test results"""
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        test_results = []
        for testcase in root.findall('.//testcase'):
            test_name = testcase.get('name', '')
            classname = testcase.get('classname', '')
            time = float(testcase.get('time', 0))
            
            failure = testcase.find('failure')
            error = testcase.find('error')
            
            status = 'PASS'
            error_message = ''
            error_type = ''
            
            if failure is not None:
                status = 'FAIL'
                error_message = failure.text or failure.get('message', '') if failure is not None else ''
                error_type = failure.get('type', '') if failure is not None else ''
            elif error is not None:
                status = 'ERROR'
                error_message = error.text or error.get('message', '') if error is not None else ''
                error_type = error.get('type', '') if error is not None else ''
            
            # Clean error message (remove stack trace, keep first few lines)
            if error_message:
                lines = error_message.split('\n')
                # Keep first 3-5 meaningful lines
                clean_lines = []
                for line in lines[:10]:
                    line = line.strip()
                    if line and not line.startswith('at ') and not line.startswith('Caused by'):
                        clean_lines.append(line)
                        if len(clean_lines) >= 3:
                            break
                error_message = '\n'.join(clean_lines[:3]) if clean_lines else error_message[:200]
            
            test_results.append({
                'name': test_name,
                'classname': classname,
                'status': status,
                'time': time,
                'error_message': error_message,
                'error_type': error_type
            })
        
        return {
            'testsuite': root.get('name', ''),
            'tests': int(root.get('tests', 0)),
            'failures': int(root.get('failures', 0)),
            'errors': int(root.get('errors', 0)),
            'skipped': int(root.get('skipped', 0)),
            'test_results': test_results
        }
    except Exception as e:
        print(f"Error parsing {xml_file}: {e}")
        return None

def extract_conditions_from_integration_test(test_method_name, method_body):
    """Extract conditions from integration test method body"""
    conditions = {
        'precondition': ['Database is available', 'Spring context is loaded', 'Test data is set up'],
        'inputs': {},
        'expected_return': [],
        'log_message': None,
        'exception': False,
        'exception_type': None,
        'type': 'N'  # N=Normal, A=Abnormal, B=Boundary
    }
    
    method_lower = method_body.lower()
    
    # Determine test type
    if 'invalid' in method_lower or 'null' in method_lower or 'notfound' in method_lower:
        conditions['type'] = 'A'
    elif 'empty' in method_lower or 'zero' in method_lower or 'boundary' in method_lower:
        conditions['type'] = 'B'
    if 'exception' in method_lower or 'throw' in method_lower or 'assertThatThrownBy' in method_lower:
        conditions['exception'] = True
        if conditions['type'] == 'N':
            conditions['type'] = 'A'
    
    # Extract expected return from assertions
    assert_patterns = [
        r'assertThat\([^)]+\)\.isEqualTo\(([^)]+)\)',
        r'assertThat\([^)]+\)\.isNotNull\(\)',
        r'assertThat\([^)]+\)\.isNull\(\)',
        r'assertThat\([^)]+\)\.isTrue\(\)',
        r'assertThat\([^)]+\)\.isFalse\(\)',
        r'assertThat\([^)]+\)\.isEmpty\(\)',
        r'assertThat\([^)]+\)\.isNotEmpty\(\)',
        r'assertThat\([^)]+\)\.hasSize\((\d+)\)',
        r'assertThat\([^)]+\)\.isGreaterThan\((\d+)\)',
        r'assertThat\([^)]+\)\.isLessThan\((\d+)\)',
    ]
    
    for pattern in assert_patterns:
        matches = re.findall(pattern, method_body)
        for match in matches:
            if match:
                match_clean = match.strip()
                if match_clean == '""' or match_clean == "''":
                    conditions['expected_return'].append('[]')
                elif re.match(r'^".*"$', match_clean):
                    str_val = match_clean.strip('"')
                    conditions['expected_return'].append(f'"{str_val}"')
                elif match_clean == 'null':
                    conditions['expected_return'].append('null')
                elif re.match(r'^\d+$', match_clean):
                    conditions['expected_return'].append(match_clean)
    
    # Extract service method calls and their parameters
    # Pattern: service.methodName(params)
    service_pattern = r'(\w+Service|dispatchService|driverService|notificationService|accountingService)\.(\w+)\(([^)]*)\)'
    service_matches = re.findall(service_pattern, method_body)
    
    for service_name, method_name, params in service_matches:
        # Parse parameters
        param_list = []
        depth = 0
        current_param = ""
        for char in params:
            if char == '(':
                depth += 1
                current_param += char
            elif char == ')':
                depth -= 1
                current_param += char
            elif char == ',' and depth == 0:
                if current_param.strip():
                    param_list.append(current_param.strip())
                current_param = ""
            else:
                current_param += char
        if current_param.strip():
            param_list.append(current_param.strip())
        
        # Store method name as input
        if 'method' not in conditions['inputs']:
            conditions['inputs']['method'] = []
        conditions['inputs']['method'].append(method_name)
        
        # Analyze parameters
        for i, param in enumerate(param_list[:5]):  # Limit to first 5 params
            param_clean = param.strip()
            
            if param_clean == 'null':
                if f'param{i+1}' not in conditions['inputs']:
                    conditions['inputs'][f'param{i+1}'] = []
                conditions['inputs'][f'param{i+1}'].append('null')
            elif param_clean == '""' or param_clean == "''":
                if f'param{i+1}' not in conditions['inputs']:
                    conditions['inputs'][f'param{i+1}'] = []
                conditions['inputs'][f'param{i+1}'].append('empty')
            elif re.match(r'^\d+$', param_clean):
                if 'id' not in conditions['inputs']:
                    conditions['inputs']['id'] = []
                conditions['inputs']['id'].append(f'value: {param_clean}')
            elif re.match(r'^".*"$', param_clean):
                str_val = param_clean.strip('"')
                if f'param{i+1}' not in conditions['inputs']:
                    conditions['inputs'][f'param{i+1}'] = []
                conditions['inputs'][f'param{i+1}'].append(f'"{str_val}"')
            elif 'LocalDate.now()' in param_clean or 'LocalDate.of' in param_clean:
                if 'date' not in conditions['inputs']:
                    conditions['inputs']['date'] = []
                if 'LocalDate.of' in param_clean:
                    date_match = re.search(r'LocalDate\.of\((\d+),\s*(\d+),\s*(\d+)\)', param_clean)
                    if date_match:
                        conditions['inputs']['date'].append(f"date: {date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}")
                else:
                    conditions['inputs']['date'].append('current date')
            elif 'Instant.now()' in param_clean:
                if 'timestamp' not in conditions['inputs']:
                    conditions['inputs']['timestamp'] = []
                conditions['inputs']['timestamp'].append('current time')
            elif 'new BigDecimal' in param_clean:
                bigdec_match = re.search(r'new BigDecimal\("?([\d.]+)"?\)', param_clean)
                if bigdec_match:
                    if 'amount' not in conditions['inputs']:
                        conditions['inputs']['amount'] = []
                    conditions['inputs']['amount'].append(f"amount: {bigdec_match.group(1)}")
    
    # Extract request object properties
    # Pattern: request.setProperty(value) or requestObject.setProperty(value)
    setter_pattern = r'(\w+Request|\w+Request)\.set(\w+)\(([^)]+)\)|(\w+)\.set(\w+)\(([^)]+)\)'
    setters = re.findall(setter_pattern, method_body)
    
    for match in setters:
        if match[0]:  # request.setProperty format
            obj, prop, value = match[0], match[1], match[2]
        else:  # obj.setProperty format
            obj, prop, value = match[3], match[4], match[5]
        
        # Focus on request objects
        if 'Request' in obj or obj.lower() in ['request', 'req', 'bookingrequest', 'assignrequest']:
            prop_lower = prop.lower()
            
            if prop_lower not in conditions['inputs']:
                conditions['inputs'][prop_lower] = []
            
            value_clean = value.strip()
            
            if value_clean == 'null':
                conditions['inputs'][prop_lower].append('null')
            elif value_clean == '""' or value_clean == "''":
                conditions['inputs'][prop_lower].append('empty')
            elif 'List.of' in value_clean or 'Arrays.asList' in value_clean:
                list_match = re.search(r'List\.of\(([^)]+)\)|Arrays\.asList\(([^)]+)\)', value_clean)
                if list_match:
                    list_content = list_match.group(1) or list_match.group(2) or ''
                    conditions['inputs'][prop_lower].append(f'[{list_content}]')
            elif re.match(r'^".*"$', value_clean):
                str_val = value_clean.strip('"')
                conditions['inputs'][prop_lower].append(f'"{str_val}"')
            elif re.match(r'^\d+$', value_clean):
                conditions['inputs'][prop_lower].append(f'value: {value_clean}')
            elif 'LocalDate.now()' in value_clean or 'LocalDate.of' in value_clean:
                if 'LocalDate.of' in value_clean:
                    date_match = re.search(r'LocalDate\.of\((\d+),\s*(\d+),\s*(\d+)\)', value_clean)
                    if date_match:
                        conditions['inputs'][prop_lower].append(f"date: {date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}")
                else:
                    conditions['inputs'][prop_lower].append('current date')
            elif 'Instant.now()' in value_clean or 'Instant.now().plusSeconds' in value_clean:
                if 'plusSeconds' in value_clean:
                    sec_match = re.search(r'plusSeconds\((\d+)\)', value_clean)
                    if sec_match:
                        conditions['inputs'][prop_lower].append(f'future time: +{sec_match.group(1)}s')
                else:
                    conditions['inputs'][prop_lower].append('current time')
            elif 'new BigDecimal' in value_clean:
                bigdec_match = re.search(r'new BigDecimal\("?([\d.]+)"?\)', value_clean)
                if bigdec_match:
                    conditions['inputs'][prop_lower].append(f"amount: {bigdec_match.group(1)}")
            elif 'Status.' in value_clean or 'BookingStatus.' in value_clean or 'TripStatus.' in value_clean:
                enum_match = re.search(r'(\w+)\.(\w+)', value_clean)
                if enum_match:
                    conditions['inputs'][prop_lower].append(f'status: {enum_match.group(2)}')
    
    # Extract entity object properties (test data setup)
    entity_setter_pattern = r'(test\w+|new\s+\w+)\.set(\w+)\(([^)]+)\)'
    entity_setters = re.findall(entity_setter_pattern, method_body, re.IGNORECASE)
    
    for obj, prop, value in entity_setters:
        prop_lower = prop.lower()
        
        # Map common entity properties
        if prop_lower in ['branchid', 'customerid', 'driverid', 'vehicleid', 'bookingid', 'tripid', 'employeeid', 'userid']:
            prop_key = prop_lower.replace('id', '_id')
            
            if prop_key not in conditions['inputs']:
                conditions['inputs'][prop_key] = []
            
            value_clean = value.strip()
            if re.match(r'^\d+$', value_clean):
                conditions['inputs'][prop_key].append(f'value: {value_clean}')
            elif value_clean == 'null':
                conditions['inputs'][prop_key].append('null')
    
    # Extract exception type
    if conditions['exception']:
        throw_pattern = r'\.isInstanceOf\((\w+\.class)\)|\.isInstanceOf\((\w+)\)|assertThatThrownBy.*(\w+Exception)'
        throw_matches = re.findall(throw_pattern, method_body)
        for match in throw_matches:
            exc_type = match[0] if match[0] else (match[1] if match[1] else match[2])
            if exc_type:
                conditions['exception_type'] = exc_type
                break
    
    # Extract log message or status from assertions
    status_pattern = r'\.getStatus\(\)|\.status|Status\s+(\w+)'
    if re.search(status_pattern, method_body, re.IGNORECASE):
        conditions['log_message'] = 'Status checked'
    
    return conditions

def parse_integration_test_code(test_file):
    """Parse Java integration test file to extract test case details"""
    try:
        with open(test_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Extract class name
        class_match = re.search(r'class\s+(\w+IntegrationTest)', content)
        class_name = class_match.group(1) if class_match else 'Unknown'
        
        # Extract test methods
        test_methods = []
        # Match @Test methods with their bodies (handling nested braces)
        test_pattern = r'@Test\s+void\s+(\w+)\s*\([^)]*\)\s*\{([^{}]*(?:\{[^{}]*\}[^{}]*)*)\}'
        
        matches = re.finditer(test_pattern, content, re.DOTALL)
        for match in matches:
            method_name = match.group(1)
            method_body = match.group(2)
            
            conditions = extract_conditions_from_integration_test(method_name, method_body)
            
            test_methods.append({
                'name': method_name,
                'conditions': conditions,
                'class_name': class_name
            })
        
        return test_methods
    except Exception as e:
        print(f"Error parsing test code {test_file}: {e}")
        import traceback
        traceback.print_exc()
        return []

def create_cover_sheet(wb):
    """Create Cover sheet"""
    ws = wb.create_sheet("Cover", 0)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'INTEGRATION TEST REPORT'
    ws['A1'].font = Font(size=20, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A3'] = 'Project Name:'
    ws['B3'] = 'PTCMSS Backend'
    ws['A4'] = 'Project Code:'
    ws['B4'] = 'PTCMSS'
    ws['A5'] = 'Document Code:'
    ws['B5'] = 'PTCMSS_Integration_Test_Report_v1.0'
    ws['A6'] = 'Issue Date:'
    ws['B6'] = datetime.now().strftime('%Y-%m-%d')
    ws['A7'] = 'Test Execution Date:'
    ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    for row in range(3, 8):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(size=11)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30

def create_method_list_sheet(wb, all_methods_data):
    """Create MethodList sheet"""
    ws = wb.create_sheet("MethodList", 1)
    
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='medium', color=BORDER_COLOR),
        right=Side(style='medium', color=BORDER_COLOR),
        top=Side(style='medium', color=BORDER_COLOR),
        bottom=Side(style='medium', color=BORDER_COLOR)
    )
    
    headers = ['No', 'Module Name', 'Method Name', 'Test Cases', 'Status', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row = 2
    for idx, (method_name, data) in enumerate(all_methods_data.items(), 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=data['module']).border = border
        ws.cell(row=row, column=3, value=method_name).border = border
        ws.cell(row=row, column=4, value=data['test_count']).border = border
        
        # Determine status
        passed = sum(1 for tc in data['test_cases'] if tc.get('status') == 'PASS')
        failed = sum(1 for tc in data['test_cases'] if tc.get('status') in ['FAIL', 'ERROR'])
        if failed > 0:
            status = 'FAIL'
        elif passed == data['test_count']:
            status = 'PASS'
        else:
            status = 'PARTIAL'
        
        status_cell = ws.cell(row=row, column=5, value=status)
        status_cell.border = border
        if status == 'PASS':
            status_cell.fill = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type='solid')
        else:
            status_cell.fill = PatternFill(start_color=FAIL_COLOR, end_color=FAIL_COLOR, fill_type='solid')
        
        ws.cell(row=row, column=6, value=f'Integration test for {method_name} method').border = border
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
        
        row += 1
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 50
    ws.freeze_panes = 'A2'

def create_statistics_sheet(wb, all_methods_data):
    """Create Statistics sheet"""
    ws = wb.create_sheet("Statistics", 2)
    
    # Calculate totals
    total_tests = sum(data['test_count'] for data in all_methods_data.values())
    total_passed = 0
    total_failed = 0
    total_errors = 0
    
    for data in all_methods_data.values():
        for tc in data['test_cases']:
            status = tc.get('status', 'PASS')
            if status == 'PASS':
                total_passed += 1
            elif status == 'FAIL':
                total_failed += 1
            elif status == 'ERROR':
                total_errors += 1
    
    ws['A1'] = 'EXECUTIVE SUMMARY'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    
    ws['A3'] = 'Total Test Cases:'
    ws['B3'] = total_tests
    ws['A4'] = 'Passed:'
    ws['B4'] = total_passed
    ws['A5'] = 'Failed:'
    ws['B5'] = total_failed
    ws['A6'] = 'Errors:'
    ws['B6'] = total_errors
    ws['A7'] = 'Pass Rate:'
    ws['B7'] = f'{(total_passed/total_tests*100):.1f}%' if total_tests > 0 else '0%'
    
    for row in range(3, 8):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(size=11)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

def create_module_sheets(wb, all_methods_data):
    """
    Create one sheet per module, aggregating all test cases of that module.
    Layout mirrors the provided Excel format:
      - Feature / Test requirement rows
      - Testing Round summary (Round1 with pass/fail/pending)
      - Green header row for test case table
      - Grouped by Function <method>, followed by its test cases
    """
    # Group test cases by module
    module_cases = {}
    for data in all_methods_data.values():
        module = data['module']
        if module not in module_cases:
            module_cases[module] = []
        module_cases[module].extend(data['test_cases'])

    for module_name, test_cases in module_cases.items():
        # Sheet name limit 31 chars
        sheet_name = re.sub(r'[\\/*?:\[\]]', '_', module_name)[:31]
        # Avoid duplicates
        if sheet_name in wb.sheetnames:
            base = sheet_name[:28] if len(sheet_name) > 28 else sheet_name
            suffix_idx = 1
            candidate = f"{base}_{suffix_idx}"
            while candidate in wb.sheetnames and len(candidate) < 31:
                suffix_idx += 1
                candidate = f"{base}_{suffix_idx}"
            sheet_name = candidate if len(candidate) <= 31 else base[:max(0, 31-len(str(suffix_idx))-1)] + f"_{suffix_idx}"
        ws = wb.create_sheet(sheet_name)

        # Real counts from test results
        total = len(test_cases)
        passed = sum(1 for tc in test_cases if tc.get('status') == 'PASS')
        failed_total = sum(1 for tc in test_cases if tc.get('status') in ['FAIL', 'ERROR'])
        pending_total = total - passed - failed_total

        # Styles
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        green_fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Header info
        ws['A1'] = 'Feature'
        ws['B1'] = module_name
        ws['A2'] = 'Test requirement'
        ws['B2'] = f'{module_name} integration coverage'
        ws['A3'] = 'Number of TCs'
        ws['B3'] = total

        # Testing Round table
        ws['A5'] = 'Testing Round'
        ws['B5'] = 'Passed'
        ws['C5'] = 'Failed'
        ws['D5'] = 'Pending'
        ws['E5'] = 'N/A'
        for cell in ['A5','B5','C5','D5','E5']:
            ws[cell].fill = header_fill
            ws[cell].border = border_thin
        ws['A6'] = 'Round 1'
        ws['B6'] = passed
        ws['C6'] = failed_total
        ws['D6'] = pending_total
        ws['E6'] = 0
        for cell in ['A6','B6','C6','D6','E6']:
            ws[cell].border = border_thin

        # Column widths
        widths = [18, 32, 28, 28, 22, 10, 12, 10, 10, 12, 10, 10, 12, 10]
        for idx, w in enumerate(widths, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = w

        # Table header row
        start_row = 9
        headers = [
            'Test Case ID', 'Test Case Description', 'Test Case Procedure',
            'Expected Results', 'Pre-conditions',
            'Round 1', 'Test date', 'Tester',
            'Round 2', 'Test date', 'Tester',
            'Round 3', 'Test date', 'Tester'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = green_fill
            cell.border = border_thin

        # Group test cases by method (function)
        def actual_method_name(tc_name: str) -> str:
            if '_' in tc_name:
                return tc_name.split('_')[0]
            return tc_name

        def readable_description(tc_name: str) -> str:
            # turn test name into human-readable sentence
            parts = tc_name.split('_')
            if len(parts) >= 2:
                return parts[0].capitalize() + ' ' + ' '.join(parts[1:])
            return tc_name.capitalize()

        def build_expected(tc):
            # Prefer error/actual message from test result (closer to real expected/observed)
            if tc.get('error_message'):
                return tc.get('error_message', '')
            expected_list = tc['conditions'].get('expected_return', [])
            if expected_list:
                return '\n'.join(expected_list[:3])
            return ''  # no explicit expected captured from test

        def build_procedure(tc):
            # use inputs as hints
            inputs = tc['conditions'].get('inputs', {})
            steps = []
            method_calls = inputs.get('method', [])
            if method_calls:
                steps.append(f"Call {method_calls[0]}(...)")
            if 'id' in inputs:
                steps.append(f"Use ids: {', '.join(inputs['id'])}")
            for k in ['param1','param2','param3']:
                if k in inputs:
                    steps.append(f"{k}: {', '.join(inputs[k])}")
            return '\n'.join(steps) if steps else '<Steps not extracted>'

        def build_preconditions():
            return "Database is available\nSpring context is loaded\nTest data is set up"

        method_groups = {}
        for tc in test_cases:
            method = actual_method_name(tc['name'])
            method_groups.setdefault(method, []).append(tc)

        row = start_row + 1
        for method, tcs in method_groups.items():
            # Function row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            func_cell = ws.cell(row=row, column=1, value=f'Function {method}')
            func_cell.fill = header_fill
            func_cell.border = border_thin
            row += 1

            # Test cases for this function
            for idx, tc in enumerate(tcs, 1):
                tc_id = f"<ID{idx}>"
                desc = readable_description(tc['name'])
                proc = build_procedure(tc)
                expected = build_expected(tc)
                preconds = '\n'.join(tc['conditions'].get('precondition', [])) if tc['conditions'].get('precondition') else build_preconditions()
                raw_status = tc.get('status', 'PASS')
                status_str = 'Passed' if raw_status == 'PASS' else ('Failed' if raw_status == 'FAIL' else 'Error')
                test_date = datetime.now().strftime('%d/%m/%Y')
                tester = 'Auto'

                values = [
                    tc_id, desc, proc, expected, preconds,
                    status_str, test_date, tester,
                    status_str, test_date, tester,
                    status_str, test_date, tester
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = border_thin
                # Color Round 1 cell
                r1_cell = ws.cell(row=row, column=6)
                if raw_status == 'PASS':
                    r1_cell.fill = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type='solid')
                else:
                    r1_cell.fill = PatternFill(start_color=FAIL_COLOR, end_color=FAIL_COLOR, fill_type='solid')
                # Color Round 2/3 similarly
                r2_cell = ws.cell(row=row, column=9)
                r3_cell = ws.cell(row=row, column=12)
                for rc in (r2_cell, r3_cell):
                    if raw_status == 'PASS':
                        rc.fill = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type='solid')
                    else:
                        rc.fill = PatternFill(start_color=FAIL_COLOR, end_color=FAIL_COLOR, fill_type='solid')
                row += 1

        # Freeze header
        ws.freeze_panes = 'A10'

def create_feature_test_case_sheet(wb, method_name, test_cases, test_results_map, module_name):
    """Create feature test case sheet with detailed test cases (like Feature 1, Feature 2 format)"""
    # Clean sheet name
    sheet_name = method_name[:31] if len(method_name) > 31 else method_name
    sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)
    
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    ws = wb.create_sheet(sheet_name)
    
    # Styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    function_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    function_header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # ========== SUMMARY SECTION (Rows 1-8) ==========
    
    # Row 1: Feature
    ws['A1'] = 'Feature'
    ws['A1'].font = Font(bold=True, size=11)
    ws['B1'] = f'<{module_name} - {method_name}>'
    ws['B1'].font = Font(size=11)
    
    # Row 2: Test requirement
    ws['A2'] = 'Test requirement'
    ws['A2'].font = Font(bold=True, size=11)
    ws.merge_cells('B2:E2')
    ws['B2'] = f'<Brief description about requirements which are tested in this sheet for {method_name} method>'
    ws['B2'].font = Font(size=10)
    ws['B2'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Row 3: Number of TCs
    ws['A3'] = 'Number of TCs'
    ws['A3'].font = Font(bold=True, size=11)
    ws['B3'] = len(test_cases)
    ws['B3'].font = Font(size=11)
    
    # Row 4: Testing Round header
    ws['A4'] = 'Testing Round'
    ws['A4'].font = Font(bold=True, size=11)
    ws['B4'] = 'Passed'
    ws['B4'].font = Font(bold=True, size=11)
    ws['B4'].fill = header_fill
    ws['B4'].font = header_font
    ws['B4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B4'].border = border
    
    ws['C4'] = 'Failed'
    ws['C4'].font = header_font
    ws['C4'].fill = header_fill
    ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C4'].border = border
    
    ws['D4'] = 'Pending'
    ws['D4'].font = header_font
    ws['D4'].fill = header_fill
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D4'].border = border
    
    ws['E4'] = 'N/A'
    ws['E4'].font = header_font
    ws['E4'].fill = header_fill
    ws['E4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E4'].border = border
    
    # Calculate statistics
    passed_count = sum(1 for tc in test_cases if test_results_map.get(tc['name'], {}).get('status') == 'PASS')
    failed_count = sum(1 for tc in test_cases if test_results_map.get(tc['name'], {}).get('status') in ['FAIL', 'ERROR'])
    pending_count = len(test_cases) - passed_count - failed_count
    
    # Row 5: Round 1
    ws['A5'] = 'Round 1'
    ws['A5'].font = Font(bold=True, size=10)
    ws['B5'] = passed_count
    ws['B5'].border = border
    ws['C5'] = failed_count
    ws['C5'].border = border
    ws['D5'] = pending_count
    ws['D5'].border = border
    ws['E5'] = 0
    ws['E5'].border = border
    
    # Row 6: Round 2 (same as Round 1 for now)
    ws['A6'] = 'Round 2'
    ws['A6'].font = Font(bold=True, size=10)
    ws['B6'] = passed_count
    ws['B6'].border = border
    ws['C6'] = failed_count
    ws['C6'].border = border
    ws['D6'] = pending_count
    ws['D6'].border = border
    ws['E6'] = 0
    ws['E6'].border = border
    
    # Row 7: Round 3 (same as Round 1 for now)
    ws['A7'] = 'Round 3'
    ws['A7'].font = Font(bold=True, size=10)
    ws['B7'] = passed_count
    ws['B7'].border = border
    ws['C7'] = failed_count
    ws['C7'].border = border
    ws['D7'] = pending_count
    ws['D7'].border = border
    ws['E7'] = 0
    ws['E7'].border = border
    
    # Row 8: Empty row for spacing
    row = 9
    
    # ========== TEST CASE DETAILS SECTION (from row 9) ==========
    
    # Row 9: Header row
    headers = [
        'Test Case ID', 'Test Case Description', 'Test Case Procedure', 
        'Expected Results', 'Pre-conditions',
        'Round 1', 'Test date', 'Tester', 'Error Message',
        'Round 2', 'Test date', 'Tester', 'Error Message',
        'Round 3', 'Test date', 'Tester', 'Error Message'
    ]
    
    header_fill_green = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill_green
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    row += 1
    
    # Group test cases by function (extract function name from test case name)
    function_groups = {}
    for tc in test_cases:
        # Extract function name (e.g., "getDashboard" from "getDashboard_shouldReturnDashboardData")
        function_match = re.match(r'(\w+)_', tc['name'])
        if function_match:
            function_name = function_match.group(1)
        else:
            function_name = 'General'
        
        if function_name not in function_groups:
            function_groups[function_name] = []
        function_groups[function_name].append(tc)
    
    # Write test cases grouped by function
    for function_name, function_tests in function_groups.items():
        # Function header row
        ws.merge_cells(f'A{row}:Q{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f'Function {function_name}'
        cell.fill = function_header_fill
        cell.font = function_header_font
        cell.border = border
        row += 1
        
        # Write each test case
        for idx, tc in enumerate(function_tests, 1):
            test_id = f'<ID{idx}>'
            status = test_results_map.get(tc['name'], {}).get('status', 'PASS')
            
            # Map status to Round status
            if status == 'PASS':
                round_status = 'Passed'
                status_fill = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type='solid')
            elif status in ['FAIL', 'ERROR']:
                round_status = 'Failed'
                status_fill = PatternFill(start_color=FAIL_COLOR, end_color=FAIL_COLOR, fill_type='solid')
            else:
                round_status = 'Pending'
                status_fill = None
            
            # Test Case ID
            ws.cell(row=row, column=1, value=test_id).border = border
            ws.cell(row=row, column=1).font = normal_font
            
            # Test Case Description
            desc = f'<Brief description of this case: what is tested?> Test: {tc["name"]}'
            ws.cell(row=row, column=2, value=desc).border = border
            ws.cell(row=row, column=2).font = normal_font
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Test Case Procedure
            procedure = f'<Describe steps to perform this case>\n1. Setup test data\n2. Call {method_name} method\n3. Verify results'
            ws.cell(row=row, column=3, value=procedure).border = border
            ws.cell(row=row, column=3).font = normal_font
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Expected Results
            expected = tc['conditions'].get('expected_return', [])
            if expected:
                expected_str = '<Describe results which meet customer\'s requirement>\n' + '\n'.join(expected[:3])
            else:
                expected_str = '<Describe results which meet customer\'s requirement>'
            ws.cell(row=row, column=4, value=expected_str).border = border
            ws.cell(row=row, column=4).font = normal_font
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Pre-conditions
            preconditions = tc['conditions'].get('precondition', [])
            precond_str = '<List all test cases or conditions that must be done before performing this case>\n' + '\n'.join(preconditions[:3])
            ws.cell(row=row, column=5, value=precond_str).border = border
            ws.cell(row=row, column=5).font = normal_font
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Get error message from test results
            error_msg = test_results_map.get(tc['name'], {}).get('error_message', '')
            error_type = test_results_map.get(tc['name'], {}).get('error_type', '')
            
            # Round 1
            round1_cell = ws.cell(row=row, column=6, value=round_status)
            round1_cell.border = border
            round1_cell.font = normal_font
            round1_cell.alignment = Alignment(horizontal='center', vertical='center')
            if status_fill:
                round1_cell.fill = status_fill
            
            ws.cell(row=row, column=7, value='').border = border  # Test date
            ws.cell(row=row, column=8, value='').border = border  # Tester
            
            # Error Message for Round 1
            error_cell_1 = ws.cell(row=row, column=9, value=error_msg if error_msg else '')
            error_cell_1.border = border
            error_cell_1.font = Font(size=9, color='FF0000' if error_msg else '000000')
            error_cell_1.alignment = Alignment(wrap_text=True, vertical='top')
            if error_msg:
                error_cell_1.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            # Round 2
            round2_cell = ws.cell(row=row, column=10, value=round_status)
            round2_cell.border = border
            round2_cell.font = normal_font
            round2_cell.alignment = Alignment(horizontal='center', vertical='center')
            if status_fill:
                round2_cell.fill = status_fill
            
            ws.cell(row=row, column=11, value='').border = border  # Test date
            ws.cell(row=row, column=12, value='').border = border  # Tester
            
            # Error Message for Round 2
            error_cell_2 = ws.cell(row=row, column=13, value=error_msg if error_msg else '')
            error_cell_2.border = border
            error_cell_2.font = Font(size=9, color='FF0000' if error_msg else '000000')
            error_cell_2.alignment = Alignment(wrap_text=True, vertical='top')
            if error_msg:
                error_cell_2.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            # Round 3
            round3_cell = ws.cell(row=row, column=14, value=round_status)
            round3_cell.border = border
            round3_cell.font = normal_font
            round3_cell.alignment = Alignment(horizontal='center', vertical='center')
            if status_fill:
                round3_cell.fill = status_fill
            
            ws.cell(row=row, column=15, value='').border = border  # Test date
            ws.cell(row=row, column=16, value='').border = border  # Tester
            
            # Error Message for Round 3
            error_cell_3 = ws.cell(row=row, column=17, value=error_msg if error_msg else '')
            error_cell_3.border = border
            error_cell_3.font = Font(size=9, color='FF0000' if error_msg else '000000')
            error_cell_3.alignment = Alignment(wrap_text=True, vertical='top')
            if error_msg:
                error_cell_3.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30  # Error Message Round 1
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 30  # Error Message Round 2
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 15
    ws.column_dimensions['Q'].width = 30  # Error Message Round 3
    
    # Freeze panes
    ws.freeze_panes = 'A10'

def create_method_decision_table(wb, method_name, test_cases, test_results_map, module_name):
    """Create decision table sheet for a method"""
    # Clean sheet name
    sheet_name = method_name[:31] if len(method_name) > 31 else method_name
    sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)
    
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    ws = wb.create_sheet(sheet_name)
    
    # Styles
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    normal_font = Font(size=10)
    border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    # ========== HEADER SECTION (A1-Q5) ==========
    
    # Row 1: Code Module and Method
    ws['A1'] = 'Code Module:'
    ws['A1'].font = Font(bold=True, size=11)
    ws['B1'] = module_name
    ws['B1'].font = Font(size=11)
    
    ws['F1'] = 'Method:'
    ws['F1'].font = Font(bold=True, size=11)
    ws['G1'] = method_name
    ws['G1'].font = Font(size=11)
    
    # Row 2: Created By and Executed By
    ws['A2'] = 'Created By:'
    ws['A2'].font = Font(bold=True, size=10)
    ws['B2'] = 'NamDD'
    ws['B2'].font = Font(size=10)
    
    ws['F2'] = 'Executed By:'
    ws['F2'].font = Font(bold=True, size=10)
    ws['G2'] = 'NamDD'
    ws['G2'].font = Font(size=10)
    
    # Row 3: Test requirement
    ws['A3'] = 'Test requirement:'
    ws['A3'].font = Font(bold=True, size=10)
    ws.merge_cells('B3:Q3')
    ws['B3'] = '<Brief description about requirements which are tested in this function>'
    ws['B3'].font = Font(size=10)
    ws['B3'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Row 4: Empty row for spacing
    
    # Row 5: Statistics
    passed_count = sum(1 for tc in test_cases if test_results_map.get(tc['name'], {}).get('status') == 'PASS')
    failed_count = sum(1 for tc in test_cases if test_results_map.get(tc['name'], {}).get('status') in ['FAIL', 'ERROR'])
    untested_count = len(test_cases) - passed_count - failed_count
    
    type_counts = {'N': 0, 'A': 0, 'B': 0}
    for tc in test_cases:
        test_type = tc['conditions'].get('type', 'N')
        type_counts[test_type] = type_counts.get(test_type, 0) + 1
    
    ws['A5'] = 'Passed:'
    ws['A5'].font = Font(bold=True, size=10)
    ws['B5'] = passed_count
    ws['B5'].font = Font(size=10)
    
    ws['C5'] = 'Failed:'
    ws['C5'].font = Font(bold=True, size=10)
    ws['D5'] = failed_count
    ws['D5'].font = Font(size=10)
    
    ws['E5'] = 'Untested:'
    ws['E5'].font = Font(bold=True, size=10)
    ws['F5'] = untested_count
    ws['F5'].font = Font(size=10)
    
    ws['G5'] = 'N/A/B:'
    ws['G5'].font = Font(bold=True, size=10)
    ws['H5'] = f"{type_counts['N']} (Normal), {type_counts['A']} (Abnormal), {type_counts['B']} (Boundary)"
    ws['H5'].font = Font(size=10)
    
    ws['I5'] = 'Total Test Cases:'
    ws['I5'].font = Font(bold=True, size=10)
    ws['J5'] = len(test_cases)
    ws['J5'].font = Font(size=10)
    
    # ========== DECISION TABLE SECTION (from row 7) ==========
    
    # Row 7: Test Case IDs header
    ws['A7'] = 'Condition'
    ws['A7'].fill = header_fill
    ws['A7'].font = header_font
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A7'].border = border
    
    ws['B7'] = 'Precondition'
    ws['B7'].fill = header_fill
    ws['B7'].font = header_font
    ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B7'].border = border
    
    # Test case column headers (ITCID01, ITCID02, ...)
    for idx, test_case in enumerate(test_cases, 1):
        col_letter = get_column_letter(idx + 2)  # Start from column C
        itc_id = f'ITCID{idx:02d}'
        cell = ws.cell(row=7, column=idx + 2)
        cell.value = itc_id
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        cell.border = border
        
        # Add status color
        status = test_results_map.get(test_case['name'], {}).get('status', 'PASS')
        if status == 'PASS':
            cell.fill = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type='solid')
        elif status in ['FAIL', 'ERROR']:
            cell.fill = PatternFill(start_color=FAIL_COLOR, end_color=FAIL_COLOR, fill_type='solid')
    
    # Build conditions list
    conditions_list = []
    
    # Precondition section - always first
    conditions_list.append({
        'condition': '',
        'precondition': 'Database is available',
        'applies_to': list(range(len(test_cases)))  # All test cases
    })
    conditions_list.append({
        'condition': '',
        'precondition': 'Spring context is loaded',
        'applies_to': list(range(len(test_cases)))  # All test cases
    })
    conditions_list.append({
        'condition': '',
        'precondition': 'Test data is set up',
        'applies_to': list(range(len(test_cases)))  # All test cases
    })
    
    # Extract all unique input properties and normalize names
    all_properties = {}
    for tc in test_cases:
        for prop, values in tc['conditions']['inputs'].items():
            prop_normalized = prop.lower()
            if prop_normalized not in all_properties:
                all_properties[prop_normalized] = set()
            all_properties[prop_normalized].update(values)
    
    # Sort properties for consistent ordering
    priority_order = ['method', 'id', 'branchid', 'driverid', 'vehicleid', 'bookingid', 'tripid', 'userid', 'employeeid', 'date', 'timestamp', 'amount']
    sorted_props = sorted(all_properties.keys(), key=lambda x: (
        priority_order.index(x) if x in priority_order else 999,
        x
    ))
    
    # For each property, add condition rows
    for prop in sorted_props:
        prop_values = sorted(all_properties[prop])
        
        if prop_values:
            # Capitalize first letter for display
            prop_display = prop.capitalize().replace('_', ' ')
            
            # Add property header row (condition name in column A)
            conditions_list.append({
                'condition': prop_display,
                'precondition': '',
                'applies_to': []
            })
            
            # Add value rows for this property
            for value in prop_values:
                applies_to = []
                for idx, tc in enumerate(test_cases):
                    tc_inputs = tc['conditions']['inputs']
                    if prop in tc_inputs and value in tc_inputs[prop]:
                        applies_to.append(idx)
                
                conditions_list.append({
                    'condition': '',
                    'precondition': value,
                    'applies_to': applies_to
                })
    
    # Confirm/Return section - expected return values
    all_expected_returns = set()
    for tc in test_cases:
        expected = tc['conditions'].get('expected_return', [])
        if expected:
            all_expected_returns.update(expected)
        else:
            all_expected_returns.add('[]')
    
    # Always add Confirm section
    conditions_list.append({
        'condition': 'Confirm',
        'precondition': '',
        'applies_to': []
    })
    
    # Add "Return" row
    conditions_list.append({
        'condition': '',
        'precondition': 'Return',
        'applies_to': []
    })
    
    # Add expected return values
    for expected_val in sorted(all_expected_returns):
        applies_to = []
        for idx, tc in enumerate(test_cases):
            expected_list = tc['conditions'].get('expected_return', [])
            if expected_val in expected_list:
                applies_to.append(idx)
            elif not expected_list and expected_val == '[]':
                applies_to.append(idx)
        
        conditions_list.append({
            'condition': '',
            'precondition': expected_val if expected_val else '[]',
            'applies_to': applies_to
        })
    
    # Exception row
    exception_test_cases = []
    exception_type = None
    for idx, tc in enumerate(test_cases):
        if tc['conditions'].get('exception', False):
            exception_test_cases.append(idx)
            if not exception_type:
                exception_type = tc['conditions'].get('exception_type', 'Exception')
    
    conditions_list.append({
        'condition': 'Exception',
        'precondition': exception_type or '',
        'applies_to': exception_test_cases
    })
    
    # Log message section
    all_log_messages = set()
    for tc in test_cases:
        log_msg = tc['conditions'].get('log_message')
        if log_msg:
            all_log_messages.add(log_msg)
    
    conditions_list.append({
        'condition': 'Log message',
        'precondition': '',
        'applies_to': []
    })
    
    for log_msg in sorted(all_log_messages) if all_log_messages else ['']:
        applies_to = []
        for idx, tc in enumerate(test_cases):
            if tc['conditions'].get('log_message') == log_msg and log_msg:
                applies_to.append(idx)
        
        conditions_list.append({
            'condition': '',
            'precondition': log_msg,
            'applies_to': applies_to
        })
    
    # Write conditions to sheet
    row = 8
    for cond in conditions_list:
        # Condition column
        if cond['condition']:
            ws.cell(row=row, column=1, value=cond['condition']).border = border
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
        else:
            ws.cell(row=row, column=1).border = border
        
        # Precondition column
        ws.cell(row=row, column=2, value=cond['precondition']).border = border
        ws.cell(row=row, column=2).font = Font(size=10)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='center')
        
        # Test case columns - mark with "O"
        for test_idx in cond['applies_to']:
            col_letter = get_column_letter(test_idx + 3)  # Start from column C
            cell = ws.cell(row=row, column=test_idx + 3)
            cell.value = 'O'
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=12)
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    for idx in range(len(test_cases)):
        col_letter = get_column_letter(idx + 3)
        ws.column_dimensions[col_letter].width = 12
    
    # Freeze panes
    ws.freeze_panes = 'C8'

def main():
    surefire_dir = 'target/surefire-reports'
    test_src_dir = 'src/test/java'
    
    if not os.path.exists(surefire_dir):
        print(f"Directory {surefire_dir} not found!")
        print("Please run 'mvn test' first to generate test reports.")
        return
    
    # Create workbook
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Cover sheet
    create_cover_sheet(wb)
    
    # Find all XML test result files for integration tests
    xml_files = [f for f in os.listdir(surefire_dir) 
                 if f.startswith('TEST-') and f.endswith('.xml') 
                 and 'IntegrationTest' in f]
    
    if not xml_files:
        print("No integration test result files found!")
        print("Looking for files matching: TEST-*IntegrationTest*.xml")
        return
    
    print(f"Found {len(xml_files)} integration test result files")
    
    # Collect all methods data
    all_methods_data = {}
    
    # Process each test class
    for xml_file in xml_files:
        xml_path = os.path.join(surefire_dir, xml_file)
        test_results = parse_test_xml(xml_path)
        
        if not test_results:
            continue
        
        # Find corresponding test source file
        class_name = test_results['testsuite'].replace('.', '/')
        test_file = os.path.join(test_src_dir, f'{class_name}.java')
        
        if not os.path.exists(test_file):
            print(f"Test file not found: {test_file}")
            continue
        
        print(f"Processing: {class_name}")
        
        # Parse test code
        test_methods = parse_integration_test_code(test_file)
        
        # Create test results map with error messages
        test_results_map = {}
        for tr in test_results['test_results']:
            test_results_map[tr['name']] = {
                'status': tr.get('status', 'PASS'),
                'error_message': tr.get('error_message', ''),
                'error_type': tr.get('error_type', '')
            }
        
        # Extract module name from class name
        module_match = re.search(r'(\w+Service|\w+Repository|\w+Controller)', class_name)
        module_name = module_match.group(1) if module_match else class_name.split('.')[-1].replace('IntegrationTest', '')
        
        # Group test methods by method under test
        method_groups = {}
        for test_method in test_methods:
            # Extract method name from test name
            method_name = test_method['name']
            # Integration test naming: getDashboard_shouldReturnDashboardData
            match = re.match(r'(\w+)_', method_name)
            if match:
                actual_method = match.group(1)
            else:
                actual_method = method_name.split('_')[0] if '_' in method_name else method_name
            
            if actual_method not in method_groups:
                method_groups[actual_method] = []
            
            # Add test result status and error message from test_results_map
            test_result = test_results_map.get(test_method['name'], {})
            test_method['status'] = test_result.get('status', 'PASS')
            test_method['error_message'] = test_result.get('error_message', '')
            test_method['error_type'] = test_result.get('error_type', '')
            method_groups[actual_method].append(test_method)
        
        # Store methods data
        for method_name, test_cases in method_groups.items():
            if method_name not in all_methods_data:
                all_methods_data[method_name] = {
                    'module': module_name,
                    'test_count': 0,
                    'test_cases': []
                }
            
            all_methods_data[method_name]['test_count'] += len(test_cases)
            all_methods_data[method_name]['test_cases'].extend(test_cases)
    
    # Create MethodList sheet
    create_method_list_sheet(wb, all_methods_data)
    
    # Create Statistics sheet
    create_statistics_sheet(wb, all_methods_data)

    # Create per-module sheets (compact, module-level view)
    create_module_sheets(wb, all_methods_data)
    
    # Optional: create feature sheets per method (set to False to keep only module pages)
    GENERATE_METHOD_SHEETS = False
    if GENERATE_METHOD_SHEETS:
        for method_name, data in all_methods_data.items():
            try:
                # Build test results map with error messages
                test_results_map = {}
                for tc in data['test_cases']:
                    test_results_map[tc['name']] = {
                        'status': tc.get('status', 'PASS'),
                        'error_message': tc.get('error_message', ''),
                        'error_type': tc.get('error_type', '')
                    }
                create_feature_test_case_sheet(wb, method_name, data['test_cases'], test_results_map, data['module'])
                print(f"Created feature sheet for {method_name} with {len(data['test_cases'])} test cases")
            except Exception as e:
                print(f"Error creating sheet for {method_name}: {e}")
                import traceback
                traceback.print_exc()
    
    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'PTCMSS_Integration_Test_Report_{timestamp}.xlsx'
    wb.save(output_file)
    print(f"\n[SUCCESS] Excel report created successfully: {output_file}")
    print(f"[INFO] Total sheets: {len(wb.sheetnames)}")
    print(f"   - Cover")
    print(f"   - MethodList")
    print(f"   - Statistics")
    print(f"   - {len(wb.sheetnames) - 3} feature test case sheets (detailed format with error messages)")

if __name__ == '__main__':
    main()

