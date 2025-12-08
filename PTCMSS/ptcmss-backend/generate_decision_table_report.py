#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Excel report with Decision Table format for test cases
Format: Cover -> MethodList -> Statistics -> Method sheets (decision tables)
"""

import xml.etree.ElementTree as ET
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
HEADER_COLOR = '366092'
PASS_COLOR = 'C6EFCE'
FAIL_COLOR = 'FFC7CE'
WARNING_COLOR = 'FFEB9C'
BORDER_COLOR = '000000'

def parse_test_xml(xml_file):
    """Parse test XML file and extract test results"""
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        test_results = []
        for testcase in root.findall('.//testcase'):
            test_name = testcase.get('name', '')
            classname = testcase.get('classname', '')
            time = float(testcase.get('time', 0))
            
            failure = testcase.find('failure')
            error = testcase.find('error')
            
            status = 'PASS'
            if failure is not None:
                status = 'FAIL'
            elif error is not None:
                status = 'ERROR'
            
            test_results.append({
                'name': test_name,
                'classname': classname,
                'status': status,
                'time': time
            })
        
        return {
            'testsuite': root.get('name', ''),
            'tests': int(root.get('tests', 0)),
            'failures': int(root.get('failures', 0)),
            'errors': int(root.get('errors', 0)),
            'skipped': int(root.get('skipped', 0)),
            'test_results': test_results
        }
    except Exception as e:
        print(f"Error parsing {xml_file}: {e}")
        return None

def extract_conditions_from_test(test_method_name, method_body):
    """Extract conditions from test method body"""
    conditions = {
        'precondition': ['Login into the system', 'Can connect to database'],
        'inputs': {},
        'expected_return': [],
        'log_message': None,
        'exception': False,
        'exception_type': None,
        'type': 'N'  # N=Normal, A=Abnormal, B=Boundary
    }
    
    method_lower = method_body.lower()
    
    # Determine test type
    if 'null' in method_lower:
        conditions['type'] = 'A'
    elif 'empty' in method_lower:
        conditions['type'] = 'B'
    if 'exception' in method_lower or 'throw' in method_lower:
        conditions['exception'] = True
        if conditions['type'] == 'N':
            conditions['type'] = 'A'
    
    # Extract expected return from assertions
    # Pattern: assertThat(result.getXXX()).isEqualTo("value")
    assert_patterns = [
        r'assertThat\([^)]+\)\.isEqualTo\(([^)]+)\)',
        r'assertThat\([^)]+\)\.isNotNull\(\)',
        r'assertThat\([^)]+\)\.isNull\(\)',
        r'assertThat\([^)]+\)\.isTrue\(\)',
        r'assertThat\([^)]+\)\.isFalse\(\)',
        r'assertThat\([^)]+\)\.isEmpty\(\)',
    ]
    
    for pattern in assert_patterns:
        matches = re.findall(pattern, method_body)
        for match in matches:
            if match:
                match_clean = match.strip()
                if match_clean == '""' or match_clean == "''":
                    conditions['expected_return'].append('[]')
                elif re.match(r'^".*"$', match_clean):
                    str_val = match_clean.strip('"')
                    conditions['expected_return'].append(f'"{str_val}"')
                elif match_clean == 'null':
                    conditions['expected_return'].append('null')
                elif re.match(r'^\d+$', match_clean):
                    conditions['expected_return'].append(match_clean)
    
    # Extract return value from when().thenReturn() - including lists
    return_pattern = r'\.thenReturn\(([^)]+)\)'
    return_matches = re.findall(return_pattern, method_body)
    for match in return_matches:
        match_clean = match.strip()
        # Check if it's a list (Arrays.asList, List.of, etc.)
        if 'Arrays.asList' in match_clean or 'List.of' in match_clean or match_clean.startswith('['):
            # Extract list values
            list_match = re.search(r'\[([^\]]*)\]|Arrays\.asList\(([^)]+)\)|List\.of\(([^)]+)\)', match_clean)
            if list_match:
                list_content = list_match.group(1) or list_match.group(2) or list_match.group(3) or ''
                if list_content.strip():
                    # Format as list
                    conditions['expected_return'].append(f'[{list_content}]')
                else:
                    conditions['expected_return'].append('[]')
        elif re.match(r'^".*"$', match_clean):
            str_val = match_clean.strip('"')
            conditions['expected_return'].append(f'"{str_val}"')
        elif match_clean == 'null':
            conditions['expected_return'].append('null')
        elif re.match(r'^\d+$', match_clean):
            conditions['expected_return'].append(match_clean)
        elif 'Optional.of' in match_clean:
            # Extract value from Optional
            opt_match = re.search(r'Optional\.of\(([^)]+)\)', match_clean)
            if opt_match:
                opt_val = opt_match.group(1).strip()
                if re.match(r'^".*"$', opt_val):
                    conditions['expected_return'].append(opt_val)
                else:
                    conditions['expected_return'].append(opt_val)
    
    # Extract log message or status code
    # Look for status codes or log patterns
    status_pattern = r'Status\s+(\d+)|status\s+(\d+)|HTTP\s+(\d+)'
    status_match = re.search(status_pattern, method_body, re.IGNORECASE)
    if status_match:
        status_code = status_match.group(1) or status_match.group(2) or status_match.group(3)
        conditions['log_message'] = f'Status {status_code}'
    
    # Extract all string literals from the test body (these are usually test values)
    string_literals = re.findall(r'"([^"]+)"', method_body)
    
    # Extract parameter modifications - focus on request objects
    # Pattern: variable.setProperty(value)
    setter_pattern = r'(\w+)\.set(\w+)\(([^)]+)\)'
    setters = re.findall(setter_pattern, method_body)
    
    # Map common request object names
    request_objects = {}
    for obj, prop, value in setters:
        # Identify request objects (LoginRequest, CreateInvoiceRequest, etc.)
        if 'Request' in obj or obj.lower() in ['request', 'req']:
            if obj not in request_objects:
                request_objects[obj] = {}
            request_objects[obj][prop] = value.strip()
        
        value_clean = value.strip()
        
        # Normalize property name (lowercase for consistency)
        prop_lower = prop.lower()
        
        # Map common property names
        if prop_lower in ['username', 'email', 'password', 'passwordhash', 'rolename', 
                          'invoicenumber', 'customername', 'amount', 'paidamount', 
                          'period', 'branchid', 'type', 'status', 'duedate', 'invoicedate']:
            prop_key = prop_lower
            
            if prop_key not in conditions['inputs']:
                conditions['inputs'][prop_key] = []
            
            if value_clean == 'null':
                conditions['inputs'][prop_key].append('null')
            elif value_clean == '""' or value_clean == "''":
                conditions['inputs'][prop_key].append('[]')
            elif 'Arrays.asList' in value_clean or 'List.of' in value_clean or value_clean.startswith('['):
                # Extract list values
                list_match = re.search(r'\[([^\]]*)\]|Arrays\.asList\(([^)]+)\)|List\.of\(([^)]+)\)', value_clean)
                if list_match:
                    list_content = list_match.group(1) or list_match.group(2) or list_match.group(3) or ''
                    if list_content.strip():
                        conditions['inputs'][prop_key].append(f'[{list_content}]')
                    else:
                        conditions['inputs'][prop_key].append('[]')
            elif re.match(r'^".*"$', value_clean):
                str_val = value_clean.strip('"')
                # Store actual value for detailed reporting
                conditions['inputs'][prop_key].append(f'"{str_val}"')
            elif re.match(r'^\d+$', value_clean):
                conditions['inputs'][prop_key].append(f'value: {value_clean}')
            elif 'LocalDate.of' in value_clean:
                date_match = re.search(r'LocalDate\.of\((\d+),\s*(\d+),\s*(\d+)\)', value_clean)
                if date_match:
                    conditions['inputs'][prop_key].append(f"date: {date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}")
            elif 'new BigDecimal' in value_clean:
                bigdec_match = re.search(r'new BigDecimal\("?(\d+)"?\)', value_clean)
                if bigdec_match:
                    conditions['inputs'][prop_key].append(f"amount: {bigdec_match.group(1)}")
            elif 'PaymentStatus.' in value_clean or 'UserStatus.' in value_clean:
                enum_match = re.search(r'(\w+)\.(\w+)', value_clean)
                if enum_match:
                    conditions['inputs'][prop_key].append(f'status: {enum_match.group(2)}')
    
    # Extract object property assignments (user.setId, user.setUsername, etc.)
    obj_setter_pattern = r'(\w+)\.set(\w+)\(([^)]+)\)'
    obj_setters = re.findall(obj_setter_pattern, method_body)
    
    for obj, prop, value in obj_setters:
        # Focus on entity objects (Users, Invoice, etc.)
        if obj.lower() in ['user', 'users', 'invoice', 'invoices', 'customer', 'booking']:
            prop_lower = prop.lower()
            if prop_lower in ['id', 'username', 'email', 'password', 'passwordhash', 'rolename']:
                prop_key = prop_lower
                
                if prop_key not in conditions['inputs']:
                    conditions['inputs'][prop_key] = []
                
                value_clean = value.strip()
                if value_clean == 'null':
                    conditions['inputs'][prop_key].append('null')
                elif re.match(r'^".*"$', value_clean):
                    str_val = value_clean.strip('"')
                    if len(str_val) >= 6 and len(str_val) <= 32:
                        conditions['inputs'][prop_key].append('6 <= && <= 32')
                    elif len(str_val) < 6 or len(str_val) > 32:
                        conditions['inputs'][prop_key].append('6 > || < 32')
                elif re.match(r'^\d+$', value_clean):
                    conditions['inputs'][prop_key].append(f'value: {value_clean}')
    
    # Extract method call parameters
    method_call_pattern = r'(\w+Service)\.(\w+)\(([^)]+)\)|exportService\.(\w+)\(([^)]+)\)'
    method_calls = re.findall(method_call_pattern, method_body)
    
    for match in method_calls:
        if match[0]:  # serviceName.methodName format
            method_name, params = match[1], match[2]
        else:  # exportService format
            method_name, params = match[3], match[4]
        
        # Parse parameters
        param_list = []
        depth = 0
        current_param = ""
        for char in params:
            if char == '(':
                depth += 1
                current_param += char
            elif char == ')':
                depth -= 1
                current_param += char
            elif char == ',' and depth == 0:
                if current_param.strip():
                    param_list.append(current_param.strip())
                current_param = ""
            else:
                current_param += char
        if current_param.strip():
            param_list.append(current_param.strip())
        
        # Analyze parameters
        for i, param in enumerate(param_list[:3]):  # Limit to first 3 params
            param_clean = param.strip()
            
            if param_clean == 'null' or param_clean == 'isNull()':
                if 'param1' not in conditions['inputs']:
                    conditions['inputs']['param1'] = []
                conditions['inputs']['param1'].append('null')
            elif param_clean == '""' or param_clean == "''":
                if 'param1' not in conditions['inputs']:
                    conditions['inputs']['param1'] = []
                conditions['inputs']['param1'].append('empty')
            elif re.match(r'^\d+$', param_clean):
                if 'id' not in conditions['inputs']:
                    conditions['inputs']['id'] = []
                conditions['inputs']['id'].append(f'value: {param_clean}')
            elif re.match(r'^".*"$', param_clean):
                str_val = param_clean.strip('"')
                if 'param1' not in conditions['inputs']:
                    conditions['inputs']['param1'] = []
                if len(str_val) >= 6 and len(str_val) <= 32:
                    conditions['inputs']['param1'].append('6 <= && <= 32')
                elif len(str_val) < 6 or len(str_val) > 32:
                    conditions['inputs']['param1'].append('6 > || < 32')
    
    # Extract exception type
    if conditions['exception']:
        throw_pattern = r'\.thenThrow\(new\s+(\w+)\(([^)]*)\)\)|isInstanceOf\((\w+\.class)\)|isInstanceOf\((\w+)\)'
        throw_matches = re.findall(throw_pattern, method_body)
        for match in throw_matches:
            exc_type = match[0] if match[0] else (match[2] if match[2] else match[3])
            if exc_type:
                conditions['exception_type'] = exc_type
                break
    
    return conditions

def parse_test_code(test_file):
    """Parse Java test file to extract test case details"""
    try:
        with open(test_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Extract class name
        class_match = re.search(r'class\s+(\w+)', content)
        class_name = class_match.group(1) if class_match else 'Unknown'
        
        # Extract test methods
        test_methods = []
        test_pattern = r'@Test\s+void\s+(\w+)\s*\([^)]*\)\s*\{([^}]+(?:\{[^}]*\}[^}]*)*)\}'
        
        matches = re.finditer(test_pattern, content, re.DOTALL)
        for match in matches:
            method_name = match.group(1)
            method_body = match.group(2)
            
            conditions = extract_conditions_from_test(method_name, method_body)
            
            test_methods.append({
                'name': method_name,
                'conditions': conditions,
                'class_name': class_name
            })
        
        return test_methods
    except Exception as e:
        print(f"Error parsing test code {test_file}: {e}")
        import traceback
        traceback.print_exc()
        return []

def create_cover_sheet(wb):
    """Create Cover sheet"""
    ws = wb.create_sheet("Cover", 0)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'UNIT TEST REPORT'
    ws['A1'].font = Font(size=20, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A3'] = 'Project Name:'
    ws['B3'] = 'PTCMSS Backend'
    ws['A4'] = 'Project Code:'
    ws['B4'] = 'PTCMSS'
    ws['A5'] = 'Document Code:'
    ws['B5'] = 'PTCMSS_Test_Report_v1.0'
    ws['A6'] = 'Issue Date:'
    ws['B6'] = datetime.now().strftime('%Y-%m-%d')
    ws['A7'] = 'Test Execution Date:'
    ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    for row in range(3, 8):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(size=11)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30

def create_method_list_sheet(wb, all_methods_data):
    """Create MethodList sheet"""
    ws = wb.create_sheet("MethodList", 1)
    
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    # Use medium border for all table cells so borders look equally bold
    border = Border(
        left=Side(style='medium', color=BORDER_COLOR),
        right=Side(style='medium', color=BORDER_COLOR),
        top=Side(style='medium', color=BORDER_COLOR),
        bottom=Side(style='medium', color=BORDER_COLOR)
    )
    
    headers = ['No', 'Module Name', 'Method Name', 'Test Cases', 'Status', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row = 2
    for idx, (method_name, data) in enumerate(all_methods_data.items(), 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=data['module']).border = border
        ws.cell(row=row, column=3, value=method_name).border = border
        ws.cell(row=row, column=4, value=data['test_count']).border = border
        
        # Determine status
        passed = sum(1 for tc in data['test_cases'] if tc.get('status') == 'PASS')
        failed = sum(1 for tc in data['test_cases'] if tc.get('status') in ['FAIL', 'ERROR'])
        if failed > 0:
            status = 'FAIL'
        elif passed == data['test_count']:
            status = 'PASS'
        else:
            status = 'PARTIAL'
        
        status_cell = ws.cell(row=row, column=5, value=status)
        status_cell.border = border
        if status == 'PASS':
            status_cell.fill = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type='solid')
        else:
            status_cell.fill = PatternFill(start_color=FAIL_COLOR, end_color=FAIL_COLOR, fill_type='solid')
        
        ws.cell(row=row, column=6, value=f'Test {method_name} method').border = border
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
        
        row += 1
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 50
    ws.freeze_panes = 'A2'

def create_statistics_sheet(wb, all_methods_data):
    """Create Statistics sheet"""
    ws = wb.create_sheet("Statistics", 2)
    
    # Calculate totals
    total_tests = sum(data['test_count'] for data in all_methods_data.values())
    total_passed = 0
    total_failed = 0
    total_errors = 0
    
    for data in all_methods_data.values():
        for tc in data['test_cases']:
            status = tc.get('status', 'PASS')
            if status == 'PASS':
                total_passed += 1
            elif status == 'FAIL':
                total_failed += 1
            elif status == 'ERROR':
                total_errors += 1
    
    ws['A1'] = 'EXECUTIVE SUMMARY'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    
    ws['A3'] = 'Total Test Cases:'
    ws['B3'] = total_tests
    ws['A4'] = 'Passed:'
    ws['B4'] = total_passed
    ws['A5'] = 'Failed:'
    ws['B5'] = total_failed
    ws['A6'] = 'Errors:'
    ws['B6'] = total_errors
    ws['A7'] = 'Pass Rate:'
    ws['B7'] = f'{(total_passed/total_tests*100):.1f}%' if total_tests > 0 else '0%'
    
    for row in range(3, 8):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(size=11)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

def create_method_decision_table(wb, method_name, test_cases, test_results_map, module_name):
    """Create decision table sheet for a method"""
    # Clean sheet name
    sheet_name = method_name[:31] if len(method_name) > 31 else method_name
    sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)
    
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    ws = wb.create_sheet(sheet_name)
    
    # Styles
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    normal_font = Font(size=10)
    border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    # ========== HEADER SECTION (A1-Q5) ==========
    
    # Row 1: Code Module and Method
    ws['A1'] = 'Code Module:'
    ws['A1'].font = Font(bold=True, size=11)
    ws['B1'] = module_name
    ws['B1'].font = Font(size=11)
    
    ws['F1'] = 'Method:'
    ws['F1'].font = Font(bold=True, size=11)
    ws['G1'] = method_name
    ws['G1'].font = Font(size=11)
    
    # Row 2: Created By and Executed By
    ws['A2'] = 'Created By:'
    ws['A2'].font = Font(bold=True, size=10)
    ws['B2'] = 'NamDD'
    ws['B2'].font = Font(size=10)
    
    ws['F2'] = 'Executed By:'
    ws['F2'].font = Font(bold=True, size=10)
    ws['G2'] = 'NamDD'
    ws['G2'].font = Font(size=10)
    
    # Row 3: Test requirement
    ws['A3'] = 'Test requirement:'
    ws['A3'].font = Font(bold=True, size=10)
    ws.merge_cells('B3:Q3')
    ws['B3'] = '<Brief description about requirements which are tested in this function>'
    ws['B3'].font = Font(size=10)
    ws['B3'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Row 4: Empty row for spacing
    
    # Row 5: Statistics
    passed_count = sum(1 for tc in test_cases if test_results_map.get(tc['name'], {}).get('status') == 'PASS')
    failed_count = sum(1 for tc in test_cases if test_results_map.get(tc['name'], {}).get('status') in ['FAIL', 'ERROR'])
    untested_count = len(test_cases) - passed_count - failed_count
    
    type_counts = {'N': 0, 'A': 0, 'B': 0}
    for tc in test_cases:
        test_type = tc['conditions'].get('type', 'N')
        type_counts[test_type] = type_counts.get(test_type, 0) + 1
    
    ws['A5'] = 'Passed:'
    ws['A5'].font = Font(bold=True, size=10)
    ws['B5'] = passed_count
    ws['B5'].font = Font(size=10)
    
    ws['C5'] = 'Failed:'
    ws['C5'].font = Font(bold=True, size=10)
    ws['D5'] = failed_count
    ws['D5'].font = Font(size=10)
    
    ws['E5'] = 'Untested:'
    ws['E5'].font = Font(bold=True, size=10)
    ws['F5'] = untested_count
    ws['F5'].font = Font(size=10)
    
    ws['G5'] = 'N/A/B:'
    ws['G5'].font = Font(bold=True, size=10)
    ws['H5'] = f"{type_counts['N']} (Normal), {type_counts['A']} (Abnormal), {type_counts['B']} (Boundary)"
    ws['H5'].font = Font(size=10)
    
    ws['I5'] = 'Total Test Cases:'
    ws['I5'].font = Font(bold=True, size=10)
    ws['J5'] = len(test_cases)
    ws['J5'].font = Font(size=10)
    
    # ========== DECISION TABLE SECTION (from row 7) ==========
    
    # Row 7: Test Case IDs header
    ws['A7'] = 'Condition'
    ws['A7'].fill = header_fill
    ws['A7'].font = header_font
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A7'].border = border
    
    ws['B7'] = 'Precondition'
    ws['B7'].fill = header_fill
    ws['B7'].font = header_font
    ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B7'].border = border
    
    # Test case column headers (UTCID01, UTCID02, ...)
    for idx, test_case in enumerate(test_cases, 1):
        col_letter = get_column_letter(idx + 2)  # Start from column C
        utc_id = f'UTCID{idx:02d}'
        cell = ws.cell(row=7, column=idx + 2)
        cell.value = utc_id
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        cell.border = border
    
    # Build conditions list
    conditions_list = []
    
    # Precondition section - always first
    conditions_list.append({
        'condition': '',
        'precondition': 'Login into the system',
        'applies_to': list(range(len(test_cases)))  # All test cases
    })
    conditions_list.append({
        'condition': '',
        'precondition': 'Can connect to database',
        'applies_to': list(range(len(test_cases)))  # All test cases
    })
    
    # Extract all unique input properties and normalize names
    all_properties = {}
    for tc in test_cases:
        for prop, values in tc['conditions']['inputs'].items():
            # Normalize property names
            prop_normalized = prop.lower()
            if prop_normalized not in all_properties:
                all_properties[prop_normalized] = set()
            all_properties[prop_normalized].update(values)
    
    # Sort properties for consistent ordering
    # Priority: username, password, email, then others
    priority_order = ['username', 'password', 'email', 'passwordhash', 'rolename', 'id']
    sorted_props = sorted(all_properties.keys(), key=lambda x: (
        priority_order.index(x) if x in priority_order else 999,
        x
    ))
    
    # For each property, add condition rows
    for prop in sorted_props:
        prop_values = sorted(all_properties[prop])
        
        if prop_values:
            # Capitalize first letter for display
            prop_display = prop.capitalize()
            
            # Add property header row (condition name in column A)
            conditions_list.append({
                'condition': prop_display,  # e.g., "Username", "Password"
                'precondition': '',
                'applies_to': []
            })
            
            # Add value rows for this property
            for value in prop_values:
                applies_to = []
                for idx, tc in enumerate(test_cases):
                    # Check if this value applies to this test case
                    tc_inputs = tc['conditions']['inputs']
                    if prop in tc_inputs and value in tc_inputs[prop]:
                        applies_to.append(idx)
                
                conditions_list.append({
                    'condition': '',  # Empty condition column
                    'precondition': value,  # Value in precondition column
                    'applies_to': applies_to
                })
    
    # Confirm/Return section - expected return values
    all_expected_returns = set()
    for tc in test_cases:
        expected = tc['conditions'].get('expected_return', [])
        if expected:
            all_expected_returns.update(expected)
        else:
            # If no expected return, add empty list
            all_expected_returns.add('[]')
    
    # Always add Confirm section
    conditions_list.append({
        'condition': 'Confirm',
        'precondition': '',
        'applies_to': []
    })
    
    # Add "Return" row
    conditions_list.append({
        'condition': '',
        'precondition': 'Return',
        'applies_to': []
    })
    
    # Add expected return values
    for expected_val in sorted(all_expected_returns):
        applies_to = []
        for idx, tc in enumerate(test_cases):
            expected_list = tc['conditions'].get('expected_return', [])
            if expected_val in expected_list:
                applies_to.append(idx)
            elif not expected_list and expected_val == '[]':
                # If test case has no expected return, mark with empty list
                applies_to.append(idx)
        
        conditions_list.append({
            'condition': '',
            'precondition': expected_val if expected_val else '[]',
            'applies_to': applies_to
        })
    
    # Exception row - always add, even if no exception
    exception_test_cases = []
    exception_type = None
    for idx, tc in enumerate(test_cases):
        if tc['conditions'].get('exception', False):
            exception_test_cases.append(idx)
            if not exception_type:
                exception_type = tc['conditions'].get('exception_type', 'Exception')
    
    conditions_list.append({
        'condition': 'Exception',
        'precondition': exception_type or '',
        'applies_to': exception_test_cases
    })
    
    # Log message section - always add header, even if empty
    all_log_messages = set()
    for tc in test_cases:
        log_msg = tc['conditions'].get('log_message')
        if log_msg:
            all_log_messages.add(log_msg)
    
    conditions_list.append({
        'condition': 'Log message',
        'precondition': '',
        'applies_to': []
    })
    
    for log_msg in sorted(all_log_messages) if all_log_messages else ['']:
        applies_to = []
        for idx, tc in enumerate(test_cases):
            if tc['conditions'].get('log_message') == log_msg and log_msg:
                applies_to.append(idx)
        
        conditions_list.append({
            'condition': '',
            'precondition': log_msg,
            'applies_to': applies_to
        })
    
    # Write conditions to sheet
    row = 8
    for cond in conditions_list:
        # Condition column
        if cond['condition']:
            ws.cell(row=row, column=1, value=cond['condition']).border = border
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
        else:
            ws.cell(row=row, column=1).border = border
        
        # Precondition column
        ws.cell(row=row, column=2, value=cond['precondition']).border = border
        ws.cell(row=row, column=2).font = Font(size=10)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='center')
        
        # Test case columns - mark with "O"
        for test_idx in cond['applies_to']:
            col_letter = get_column_letter(test_idx + 3)  # Start from column C
            cell = ws.cell(row=row, column=test_idx + 3)
            cell.value = 'O'
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=12)
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    for idx in range(len(test_cases)):
        col_letter = get_column_letter(idx + 3)
        ws.column_dimensions[col_letter].width = 12
    
    # Freeze panes
    ws.freeze_panes = 'C8'

def main():
    surefire_dir = 'target/surefire-reports'
    test_src_dir = 'src/test/java'
    
    if not os.path.exists(surefire_dir):
        print(f"Directory {surefire_dir} not found!")
        return
    
    # Create workbook
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Cover sheet
    create_cover_sheet(wb)
    
    # Find all XML test result files
    xml_files = [f for f in os.listdir(surefire_dir) if f.startswith('TEST-') and f.endswith('.xml')]
    
    print(f"Found {len(xml_files)} test result files")
    
    # Collect all methods data
    all_methods_data = {}
    
    # Process each test class
    for xml_file in xml_files:
        xml_path = os.path.join(surefire_dir, xml_file)
        test_results = parse_test_xml(xml_path)
        
        if not test_results:
            continue
        
        # Find corresponding test source file
        class_name = test_results['testsuite'].replace('.', '/')
        test_file = os.path.join(test_src_dir, f'{class_name}.java')
        
        if not os.path.exists(test_file):
            print(f"Test file not found: {test_file}")
            continue
        
        # Parse test code
        test_methods = parse_test_code(test_file)
        
        # Create test results map
        test_results_map = {tr['name']: tr for tr in test_results['test_results']}
        
        # Extract module name from class name
        module_match = re.search(r'(\w+Service|\w+Repository|\w+Controller)', class_name)
        module_name = module_match.group(1) if module_match else class_name.split('.')[-1].replace('Test', '')
        
        # Group test methods by method under test
        method_groups = {}
        for test_method in test_methods:
            # Extract method name from test name
            method_name = test_method['name']
            match = re.match(r'(\w+)_when', method_name)
            if match:
                actual_method = match.group(1)
            else:
                actual_method = method_name.split('_')[0] if '_' in method_name else method_name
            
            if actual_method not in method_groups:
                method_groups[actual_method] = []
            
            # Add test result status
            test_method['status'] = test_results_map.get(test_method['name'], {}).get('status', 'PASS')
            method_groups[actual_method].append(test_method)
        
        # Store methods data
        for method_name, test_cases in method_groups.items():
            if method_name not in all_methods_data:
                all_methods_data[method_name] = {
                    'module': module_name,
                    'test_count': 0,
                    'test_cases': []
                }
            
            all_methods_data[method_name]['test_count'] += len(test_cases)
            all_methods_data[method_name]['test_cases'].extend(test_cases)
    
    # Create MethodList sheet
    create_method_list_sheet(wb, all_methods_data)
    
    # Create Statistics sheet
    create_statistics_sheet(wb, all_methods_data)
    
    # Create decision table sheets for each method
    for method_name, data in all_methods_data.items():
        try:
            test_results_map = {tc['name']: {'status': tc.get('status', 'PASS')} for tc in data['test_cases']}
            create_method_decision_table(wb, method_name, data['test_cases'], test_results_map, data['module'])
            print(f"Created decision table for {method_name} with {len(data['test_cases'])} test cases")
        except Exception as e:
            print(f"Error creating sheet for {method_name}: {e}")
            import traceback
            traceback.print_exc()
    
    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'PTCMSS_Decision_Table_Test_Report_{timestamp}.xlsx'
    wb.save(output_file)
    print(f"\nExcel report created successfully: {output_file}")
    print(f"Total sheets: {len(wb.sheetnames)}")

if __name__ == '__main__':
    main()
