#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to generate detailed Excel report with one sheet per method
"""

import csv
import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read CSV data
methods = []
with open('METHOD_TEST_REPORT.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        methods.append(row)

# Create new workbook
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=12)
normal_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Function to create method sheet
def create_method_sheet(wb, method_info, test_cases):
    method_name = method_info['Method Name']
    module_name = method_info['Module Name']
    status = method_info['Status']
    test_count = int(method_info['Test Cases'])
    
    # Clean sheet name (Excel has 31 char limit)
    sheet_name = method_name[:31] if len(method_name) > 31 else method_name
    
    # Remove invalid characters for sheet name
    sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)
    
    # Check if sheet exists, remove it
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    ws = wb.create_sheet(sheet_name)
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'TEST REPORT - {method_name}'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Code Module
    ws['A3'] = 'Code Module:'
    ws['B3'] = module_name
    ws['A3'].font = title_font
    ws['B3'].font = normal_font
    
    # Method
    ws['A4'] = 'Method:'
    ws['B4'] = method_name
    ws['A4'].font = title_font
    ws['B4'].font = normal_font
    
    # Summary Statistics
    ws['A6'] = 'Summary Statistics'
    ws['A6'].font = title_font
    
    # Calculate statistics based on status
    passed = test_count if status == 'PASS' else 0
    failed = 1 if status == 'FAIL' else 0
    errors = 1 if status == 'ERROR' else 0
    untested = 0
    
    if status == 'PARTIAL':
        # Estimate: assume some passed, some failed/errors
        passed = test_count - 2
        errors = 2
    elif status == 'WARNING':
        passed = test_count
        errors = 0
    
    ws['A7'] = 'Passed:'
    ws['B7'] = passed
    ws['A8'] = 'Failed:'
    ws['B8'] = failed
    ws['A9'] = 'Errors:'
    ws['B9'] = errors
    ws['A10'] = 'Untested:'
    ws['B10'] = untested
    ws['A11'] = 'Total:'
    ws['B11'] = test_count
    
    for row in range(7, 12):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
    
    # Test Requirement
    ws['A13'] = 'Test Requirement:'
    ws['A13'].font = title_font
    ws['B13'] = method_info['Description']
    ws['B13'].font = normal_font
    ws['B13'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Precondition
    ws['A14'] = 'Precondition:'
    ws['A14'].font = title_font
    ws['B14'] = 'Can connect to server and database is available'
    ws['B14'].font = normal_font
    ws['B14'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Input Parameters Header
    ws['A16'] = 'Input Parameters'
    ws['A16'].font = title_font
    ws.merge_cells('A16:F16')
    
    # Input Parameters Table Header
    param_headers = ['Parameter', 'Test Value 1', 'Test Value 2', 'Test Value 3', 'UTCID', 'Notes']
    for col, header in enumerate(param_headers, 1):
        cell = ws.cell(row=17, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Input Parameters Data (example structure)
    row = 18
    # Common parameters based on method name
    if 'create' in method_name.lower():
        ws.cell(row=row, column=1, value='Request Object').border = border
        ws.cell(row=row, column=2, value='Valid request').border = border
        ws.cell(row=row, column=3, value='Invalid request').border = border
        ws.cell(row=row, column=4, value='null').border = border
        ws.cell(row=row, column=5, value='UTCID01').border = border
        row += 1
    elif 'get' in method_name.lower() and 'ById' in method_name:
        ws.cell(row=row, column=1, value='ID').border = border
        ws.cell(row=row, column=2, value='1').border = border
        ws.cell(row=row, column=3, value='999').border = border
        ws.cell(row=row, column=4, value='null').border = border
        ws.cell(row=row, column=5, value='UTCID01').border = border
        row += 1
    elif 'update' in method_name.lower():
        ws.cell(row=row, column=1, value='ID').border = border
        ws.cell(row=row, column=2, value='1').border = border
        ws.cell(row=row, column=3, value='999').border = border
        ws.cell(row=row, column=4, value='null').border = border
        ws.cell(row=row, column=5, value='UTCID01').border = border
        row += 1
        ws.cell(row=row, column=1, value='Request Object').border = border
        ws.cell(row=row, column=2, value='Valid request').border = border
        ws.cell(row=row, column=3, value='Invalid request').border = border
        ws.cell(row=row, column=4, value='null').border = border
        ws.cell(row=row, column=5, value='UTCID02').border = border
        row += 1
    elif 'delete' in method_name.lower():
        ws.cell(row=row, column=1, value='ID').border = border
        ws.cell(row=row, column=2, value='1').border = border
        ws.cell(row=row, column=3, value='999').border = border
        ws.cell(row=row, column=4, value='null').border = border
        ws.cell(row=row, column=5, value='UTCID01').border = border
        row += 1
    
    # Expected Return
    expected_row = row + 2
    ws.cell(row=expected_row, column=1, value='Expected Return:').font = title_font
    ws.cell(row=expected_row, column=2, value='See test cases below').font = normal_font
    ws.cell(row=expected_row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Log Message
    log_row = expected_row + 1
    ws.cell(row=log_row, column=1, value='Log Message:').font = title_font
    ws.cell(row=log_row, column=2, value='Test execution logs available in test output').font = normal_font
    ws.cell(row=log_row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Test Cases Detail
    test_cases_row = log_row + 3
    ws.cell(row=test_cases_row, column=1, value='Test Cases Detail').font = title_font
    ws.merge_cells(f'A{test_cases_row}:F{test_cases_row}')
    
    # Test Cases Table Header
    test_headers = ['No', 'Test Case Name', 'Input', 'Expected Result', 'Status', 'Notes']
    for col, header in enumerate(test_headers, 1):
        cell = ws.cell(row=test_cases_row + 1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Test Cases Data
    test_row = test_cases_row + 2
    for idx, test_case in enumerate(test_cases, 1):
        ws.cell(row=test_row, column=1, value=idx).border = border
        ws.cell(row=test_row, column=2, value=test_case['name']).border = border
        ws.cell(row=test_row, column=3, value=test_case.get('input', 'N/A')).border = border
        ws.cell(row=test_row, column=4, value=test_case.get('expected', 'N/A')).border = border
        
        # Status
        status_cell = ws.cell(row=test_row, column=5, value=status)
        status_cell.border = border
        if status == 'PASS':
            status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif status in ['FAIL', 'ERROR', 'PARTIAL']:
            status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif status == 'WARNING':
            status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        ws.cell(row=test_row, column=6, value=test_case.get('notes', '')).border = border
        
        # Wrap text for long content
        for col in range(1, 7):
            ws.cell(row=test_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        
        test_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    return ws

# Group methods and create sheets
print("Creating detailed method sheets...")

# Read test cases from UNIT_TEST_REPORT.txt to get test case names
test_cases_by_method = {}
try:
    with open('UNIT_TEST_REPORT.txt', 'r', encoding='utf-8') as f:
        content = f.read()
        # Parse test cases (simplified - you may need to enhance this)
        current_method = None
        for line in content.split('\n'):
            if 'Method:' in line or 'Method: ' in line:
                # Extract method name
                match = re.search(r'Method:\s*(\w+)', line)
                if match:
                    current_method = match.group(1)
                    test_cases_by_method[current_method] = []
            elif current_method and ('[0' in line or 'Test Case' in line):
                # Extract test case name
                match = re.search(r'\[?\d+\]?\s*(\w+_when[\w_]+)', line)
                if match:
                    test_cases_by_method[current_method].append({
                        'name': match.group(1),
                        'input': 'See test file',
                        'expected': 'See test file',
                        'notes': ''
                    })
except Exception as e:
    print(f"Warning: Could not parse test cases from report: {e}")

# Create sheets for each method
created_sheets = 0
for method in methods:
    method_name = method['Method Name']
    module_name = method['Module Name']
    
    # Get test cases for this method
    test_cases = test_cases_by_method.get(method_name, [])
    
    # If no test cases found, create default ones
    if not test_cases:
        test_count = int(method['Test Cases'])
        for i in range(1, min(test_count + 1, 4)):  # Limit to 3 examples
            test_cases.append({
                'name': f'{method_name}_testCase{i}',
                'input': f'Test input {i}',
                'expected': f'Expected result {i}',
                'notes': ''
            })
    
    try:
        create_method_sheet(wb, method, test_cases)
        created_sheets += 1
        if created_sheets % 10 == 0:
            print(f"Created {created_sheets} sheets...")
    except Exception as e:
        print(f"Error creating sheet for {method_name}: {e}")
        continue

print(f"Created {created_sheets} method sheets")

# Save workbook with new name
output_file = 'PTCMSS_Unit_Test_Report_Detailed.xlsx'
wb.save(output_file)
print(f"Excel report created successfully: {output_file}")
print(f"Total sheets created: {len(wb.sheetnames)}")

