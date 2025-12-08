#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to generate Excel report for Unit Test Results
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Read CSV data
methods = []
with open('METHOD_TEST_REPORT.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        methods.append(row)

# Create workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ==================== COVER SHEET ====================
cover = wb.create_sheet("Cover", 0)
cover.merge_cells('A1:F1')
cover['A1'] = 'UNIT TEST REPORT'
cover['A1'].font = Font(size=20, bold=True)
cover['A1'].alignment = Alignment(horizontal='center', vertical='center')

cover['A3'] = 'Project Name:'
cover['B3'] = 'PTCMSS Backend'
cover['A4'] = 'Project Code:'
cover['B4'] = 'PTCMSS'
cover['A5'] = 'Document Code:'
cover['B5'] = 'PTCMSS_Test_Report_v1.0'
cover['A6'] = 'Issue Date:'
cover['B6'] = '2025-12-08'
cover['A7'] = 'Test Execution Date:'
cover['B7'] = '2025-12-08 04:37:35'

# Format cover sheet
for row in range(3, 8):
    cover[f'A{row}'].font = Font(bold=True)
    cover[f'B{row}'].font = Font(size=11)

cover.column_dimensions['A'].width = 20
cover.column_dimensions['B'].width = 30

# ==================== METHOD LIST SHEET ====================
method_list = wb.create_sheet("MethodList", 1)

# Header
headers = ['No', 'Module Name', 'Method Name', 'Test Cases', 'Status', 'Description']
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
for col, header in enumerate(headers, 1):
    cell = method_list.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Write data
for idx, method in enumerate(methods, 2):
    method_list.cell(row=idx, column=1, value=int(method['No']))
    method_list.cell(row=idx, column=2, value=method['Module Name'])
    method_list.cell(row=idx, column=3, value=method['Method Name'])
    method_list.cell(row=idx, column=4, value=int(method['Test Cases']))
    method_list.cell(row=idx, column=5, value=method['Status'])
    method_list.cell(row=idx, column=6, value=method['Description'])
    
    # Apply formatting
    for col in range(1, 7):
        cell = method_list.cell(row=idx, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Color code status
        if method['Status'] == 'PASS':
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif method['Status'] == 'FAIL':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif method['Status'] == 'ERROR':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        elif method['Status'] == 'PARTIAL':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif method['Status'] == 'WARNING':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        elif method['Status'] == 'SKIP':
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

# Set column widths
method_list.column_dimensions['A'].width = 6
method_list.column_dimensions['B'].width = 25
method_list.column_dimensions['C'].width = 35
method_list.column_dimensions['D'].width = 12
method_list.column_dimensions['E'].width = 12
method_list.column_dimensions['F'].width = 50

# Freeze header row
method_list.freeze_panes = 'A2'

# ==================== STATISTICS SHEET ====================
stats = wb.create_sheet("Statistics", 2)

# Executive Summary
stats['A1'] = 'EXECUTIVE SUMMARY'
stats['A1'].font = Font(size=16, bold=True)
stats.merge_cells('A1:B1')

stats['A3'] = 'Total Test Cases:'
stats['B3'] = 451
stats['A4'] = 'Passed:'
stats['B4'] = 415
stats['A5'] = 'Failed:'
stats['B5'] = 8
stats['A6'] = 'Errors:'
stats['B6'] = 26
stats['A7'] = 'Skipped:'
stats['B7'] = 2
stats['A8'] = 'Pass Rate:'
stats['B8'] = '92.0%'

for row in range(3, 9):
    stats[f'A{row}'].font = Font(bold=True)
    stats[f'B{row}'].font = Font(size=11)

# Module Statistics
stats['A10'] = 'MODULE STATISTICS'
stats['A10'].font = Font(size=16, bold=True)
stats.merge_cells('A10:F10')

# Module headers
module_headers = ['No', 'Module Name', 'Passed', 'Failed', 'Errors', 'Total']
for col, header in enumerate(module_headers, 1):
    cell = stats.cell(row=12, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Module data
module_data = [
    [1, 'InvoiceService', 77, 0, 0, 77],
    [2, 'CustomerService', 25, 0, 0, 25],
    [3, 'DriverService', 27, 0, 0, 27],
    [4, 'BookingService', 23, 1, 0, 24],
    [5, 'EmployeeService', 21, 0, 2, 23],
    [6, 'ExpenseRequestService', 22, 0, 0, 22],
    [7, 'DebtService', 18, 0, 0, 18],
    [8, 'AccountingService', 17, 0, 0, 17],
    [9, 'AuthenticationService', 16, 0, 2, 18],
    [10, 'UserService', 16, 0, 2, 18],
    [11, 'RoleService', 18, 0, 0, 18],
    [12, 'JwtService', 13, 0, 0, 13],
    [13, 'SystemSettingService', 13, 0, 0, 13],
    [14, 'RatingService', 11, 0, 0, 11],
    [15, 'ApprovalSyncService', 10, 0, 0, 10],
    [16, 'PaymentService', 9, 0, 0, 9],
    [17, 'PasswordService', 9, 0, 0, 9],
    [18, 'VehicleCategoryService', 8, 0, 0, 8],
    [19, 'AnalyticsService', 6, 2, 10, 18],
    [20, 'DispatchService', 5, 1, 0, 6],
    [21, 'LocalImageService', 5, 0, 0, 5],
    [22, 'ApprovalService', 4, 0, 0, 4],
    [23, 'VehicleService', 4, 0, 0, 4],
    [24, 'BranchService', 15, 0, 1, 16],
    [25, 'DepositService', 14, 0, 4, 18],
    [26, 'NotificationService', 13, 0, 5, 18],
    [27, 'BookingVehicleDetailsRepository', 1, 0, 0, 1],
    [28, 'PtcmssBackendApplication', 0, 0, 1, 1],
    ['', 'TOTAL', 415, 8, 26, 451]
]

for idx, row_data in enumerate(module_data, 13):
    for col, value in enumerate(row_data, 1):
        cell = stats.cell(row=idx, column=col)
        cell.value = value
        cell.border = border
        if idx == len(module_data) + 12:  # Total row
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# Set column widths
stats.column_dimensions['A'].width = 6
stats.column_dimensions['B'].width = 35
stats.column_dimensions['C'].width = 10
stats.column_dimensions['D'].width = 10
stats.column_dimensions['E'].width = 10
stats.column_dimensions['F'].width = 10

# Method Statistics Summary
stats['A42'] = 'METHOD STATISTICS SUMMARY'
stats['A42'].font = Font(size=16, bold=True)
stats.merge_cells('A42:B42')

stats['A44'] = 'Total Methods:'
stats['B44'] = 204
stats['A45'] = 'Methods with Tests:'
stats['B45'] = 203
stats['A46'] = 'Methods without Tests:'
stats['B46'] = 1

stats['A48'] = 'Status Breakdown:'
stats['A48'].font = Font(bold=True)
stats['A49'] = '- PASS:'
stats['B49'] = '180 methods (88.2%)'
stats['A50'] = '- PARTIAL:'
stats['B50'] = '4 methods (2.0%)'
stats['A51'] = '- ERROR:'
stats['B51'] = '6 methods (2.9%)'
stats['A52'] = '- FAIL:'
stats['B52'] = '2 methods (1.0%)'
stats['A53'] = '- WARNING:'
stats['B53'] = '11 methods (5.4%)'
stats['A54'] = '- SKIP:'
stats['B54'] = '1 method (0.5%)'

for row in range(44, 55):
    stats[f'A{row}'].font = Font(bold=True)

# ==================== SAVE FILE ====================
filename = 'PTCMSS_Unit_Test_Report.xlsx'
wb.save(filename)
print(f"Excel report generated successfully: {filename}")

